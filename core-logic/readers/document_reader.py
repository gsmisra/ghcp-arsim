from __future__ import annotations

import csv
import zipfile
from pathlib import Path
from typing import Any, Iterator

from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from pypdf import PdfReader

from models.requirement import RequirementArtifact


class DocumentReader:
    SUPPORTED_TYPES = {"docx", "xlsx", "pdf", "csv"}

    def read(self, file_path: str, options: dict[str, Any] | None = None) -> RequirementArtifact:
        path = Path(file_path)
        ext = path.suffix.lower().replace(".", "")
        options = options or {}

        if ext not in self.SUPPORTED_TYPES:
            raise ValueError(f"Unsupported file type: {ext}")

        raw_text = {
            "docx": self._read_docx,
            "xlsx": self._read_xlsx,
            "pdf": self._read_pdf,
            "csv": self._read_csv,
        }[ext](path, options)

        return RequirementArtifact(
            source_type="document",
            title=path.stem,
            raw_text=raw_text,
            metadata={"file_path": str(path), "file_type": ext, "parse_options": options},
        )

    def get_sheet_names(self, file_path: str) -> list[str]:
        path = Path(file_path)
        ext = path.suffix.lower().replace(".", "")
        if ext != "xlsx":
            return []
        workbook = load_workbook(path, data_only=True, read_only=True)
        return list(workbook.sheetnames)

    def _read_docx(self, path: Path, options: dict[str, Any] | None = None) -> str:
        options = options or {}
        if not zipfile.is_zipfile(path):
            raise ValueError("Invalid .docx file: the uploaded file is not a valid Word document archive.")

        try:
            doc = Document(path)
        except Exception as error:
            raise ValueError(f"Unable to read .docx file: {error}") from error

        page_start = int(options.get("page_start", 1) or 1)
        page_end = int(options.get("page_end", 0) or 0)
        if page_start < 1:
            raise ValueError("DOCX page_start must be >= 1")

        pages: list[list[str]] = [[]]
        all_blocks: list[str] = []
        has_explicit_page_break = False
        for block in self._iter_docx_blocks(doc):
            text = self._docx_block_to_text(block)
            if text:
                all_blocks.append(text)
                pages[-1].append(text)

            block_xml = block._element.xml
            page_break_before = "pageBreakBefore" in block_xml
            page_break_after = 'w:type="page"' in block_xml
            rendered_page_break = "lastRenderedPageBreak" in block_xml
            if page_break_before or page_break_after:
                has_explicit_page_break = True
                pages.append([])
            elif rendered_page_break:
                # Word may persist rendered page breaks without explicit manual breaks.
                has_explicit_page_break = True
                pages.append([])

        if has_explicit_page_break:
            content_pages = ["\n".join(block for block in page if block).strip() for page in pages]
            content_pages = [page for page in content_pages if page]
        else:
            # Many DOCX files do not store true rendered page boundaries.
            # Fallback: deterministic pagination by character budget to keep
            # page ranges meaningful for preview/generation.
            chars_per_page = int(options.get("docx_chars_per_page", 3500) or 3500)
            chars_per_page = max(500, chars_per_page)
            content_pages = []
            current_page: list[str] = []
            current_chars = 0
            for block in all_blocks:
                block_chars = max(1, len(block))
                if current_page and current_chars + block_chars > chars_per_page:
                    content_pages.append("\n".join(current_page).strip())
                    current_page = [block]
                    current_chars = block_chars
                else:
                    current_page.append(block)
                    current_chars += block_chars

            if current_page:
                content_pages.append("\n".join(current_page).strip())
            content_pages = [page for page in content_pages if page]

        if not content_pages:
            return ""

        total_pages = len(content_pages)
        effective_page_start = min(page_start, total_pages)
        effective_page_end = min(page_end or total_pages, total_pages)
        if effective_page_end < effective_page_start:
            effective_page_end = effective_page_start

        sliced_pages = content_pages[effective_page_start - 1 : effective_page_end]
        return "\n\n".join(
            f"## Page {index + effective_page_start}\n{page}" for index, page in enumerate(sliced_pages)
        )

    def _iter_docx_blocks(self, doc: Document) -> Iterator[Paragraph | Table]:
        body = doc._element.body
        for child in body.iterchildren():
            if child.tag.endswith("}p"):
                yield Paragraph(child, doc)
            elif child.tag.endswith("}tbl"):
                yield Table(child, doc)

    def _docx_block_to_text(self, block: Paragraph | Table) -> str:
        if isinstance(block, Paragraph):
            return block.text.strip()

        rows: list[str] = []
        for row in block.rows:
            cols = [" ".join(cell.text.split()).strip() for cell in row.cells]
            if any(col for col in cols):
                rows.append(" | ".join(cols))
        return "\n".join(rows).strip()

    def _read_xlsx(self, path: Path, options: dict[str, Any] | None = None) -> str:
        options = options or {}
        workbook = load_workbook(path, data_only=True)
        selected_sheets = options.get("sheet_names") or workbook.sheetnames
        row_start = int(options.get("row_start", 1) or 1)
        row_end = int(options.get("row_end", 0) or 0)

        if row_start < 1:
            raise ValueError("Excel row_start must be >= 1")
        if row_end and row_end < row_start:
            raise ValueError("Excel row_end must be greater than or equal to row_start")

        content = []
        for sheet_name in selected_sheets:
            if sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            content.append(f"## Sheet: {sheet_name}")

            rows = []
            for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if index < row_start:
                    continue
                if row_end and index > row_end:
                    break

                normalized = ["" if cell is None else str(cell) for cell in row]
                if any(col.strip() for col in normalized):
                    rows.append(normalized)

            if not rows:
                content.append("(empty sheet)")
                continue

            csv_lines = []
            for row in rows:
                escaped = [f'"{col.replace("\"", "\"\"")}"' for col in row]
                csv_lines.append(",".join(escaped))
            content.append("\n".join(csv_lines))

        return "\n".join(content)

    def _read_pdf(self, path: Path, options: dict[str, Any] | None = None) -> str:
        options = options or {}
        reader = PdfReader(str(path))
        pages = []
        total_pages = len(reader.pages)
        page_start = int(options.get("page_start", 1) or 1)
        page_end = int(options.get("page_end", total_pages) or total_pages)

        if page_start < 1:
            raise ValueError("PDF page_start must be >= 1")
        if page_start > total_pages:
            raise ValueError(f"PDF page_start ({page_start}) exceeds total pages ({total_pages})")
        if page_end < page_start:
            raise ValueError("PDF page_end must be greater than or equal to page_start")

        page_end = min(page_end, total_pages)

        for idx, page in enumerate(reader.pages, start=1):
            if idx < page_start or idx > page_end:
                continue
            page_text = page.extract_text() or ""
            pages.append(f"## Page {idx}\n{page_text.strip()}")
        return "\n\n".join(pages)

    def _read_csv(self, path: Path, options: dict[str, Any] | None = None) -> str:
        options = options or {}
        row_start = int(options.get("row_start", 1) or 1)
        row_end = int(options.get("row_end", 0) or 0)
        if row_start < 1:
            raise ValueError("CSV row_start must be >= 1")
        if row_end and row_end < row_start:
            raise ValueError("CSV row_end must be greater than or equal to row_start")

        rows = []
        with path.open("r", encoding="utf-8-sig", newline="") as csv_file:
            reader = csv.reader(csv_file)
            for index, row in enumerate(reader, start=1):
                if index < row_start:
                    continue
                if row_end and index > row_end:
                    break
                rows.append(["" if col is None else str(col) for col in row])

        csv_lines = []
        for row in rows:
            escaped = [f'"{col.replace("\"", "\"\"")}"' for col in row]
            csv_lines.append(",".join(escaped))
        return "\n".join(csv_lines)
